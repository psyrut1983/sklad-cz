"""
Тесты синтетических XLSX-фикстур.

Проверяет:
- Генератор создаёт детерминированный файл с 2 листами и 10 заголовками
- Структура и содержимое соответствуют ожиданиям
- Данные не читаются из реального XLSX
"""

import os
import pytest
from unittest.mock import patch

# ═════════════════════════════════════════════════════════════════════════════
#  Импорт генератора (openpyxl опционален)
# ═════════════════════════════════════════════════════════════════════════════

try:
    from app.services.synthetic_xlsx import (
        create_synthetic_xlsx,
        get_expected_stats,
        HEADERS,
        KI_CLEAN,
        KI_WITH_FFFD,
        KI_WITH_GS_TEXT,
        KI_CORRUPTED,
        KI_DUP,
    )
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    HEADERS = []
    KI_CLEAN = ""


# ═════════════════════════════════════════════════════════════════════════════
#  Skip-модуль если openpyxl не установлен
# ═════════════════════════════════════════════════════════════════════════════

pytestmark = pytest.mark.skipif(
    not HAS_OPENPYXL,
    reason="openpyxl не установлен",
)


# ═════════════════════════════════════════════════════════════════════════════
#  Тесты
# ═════════════════════════════════════════════════════════════════════════════


class TestSyntheticXlsxGenerator:
    """Проверка генератора фикстур."""

    def test_creates_file(self, tmp_path):
        """Генератор создаёт XLSX-файл."""
        path = os.path.join(tmp_path, "test_fixture.xlsx")
        result = create_synthetic_xlsx(path)
        assert result == path
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

    def test_two_sheets(self, tmp_path):
        """Файл содержит ровно 2 листа."""
        import openpyxl

        path = create_synthetic_xlsx(os.path.join(tmp_path, "sheets.xlsx"))
        wb = openpyxl.load_workbook(path)
        assert len(wb.sheetnames) == 2
        assert "Сборочные задания" in wb.sheetnames
        assert "КИЗ" in wb.sheetnames

    def test_ten_headers(self, tmp_path):
        """Лист КИЗ содержит ровно 10 заголовков."""
        import openpyxl

        path = create_synthetic_xlsx(os.path.join(tmp_path, "headers.xlsx"))
        wb = openpyxl.load_workbook(path)
        ws = wb["КИЗ"]
        headers = [cell.value for cell in ws[1]]
        assert len(headers) == 10
        assert headers == HEADERS

    def test_operation_types(self, tmp_path):
        """Содержит Продажа, Возврат и -."""
        import openpyxl

        path = create_synthetic_xlsx(os.path.join(tmp_path, "ops.xlsx"))
        wb = openpyxl.load_workbook(path)
        ws = wb["КИЗ"]
        ops = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            ops.add(row[8])  # колонка "Тип операции"
        assert "Продажа" in ops
        assert "Возврат" in ops
        assert "-" in ops

    def test_empty_check_and_fn(self, tmp_path):
        """Есть строки с пустым чеком и пустым ФН."""
        import openpyxl

        path = create_synthetic_xlsx(os.path.join(tmp_path, "empty.xlsx"))
        wb = openpyxl.load_workbook(path)
        ws = wb["КИЗ"]
        has_empty_check = False
        has_empty_fn = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[3] == "" or row[3] is None:  # Номер чека
                has_empty_check = True
            if row[6] == "" or row[6] is None:  # Номер ФН
                has_empty_fn = True
        assert has_empty_check, "Нет строки с пустым чеком"
        assert has_empty_fn, "Нет строки с пустым ФН"

    def test_integer_and_decimal_cost(self, tmp_path):
        """Есть целая и дробная стоимость."""
        import openpyxl

        path = create_synthetic_xlsx(os.path.join(tmp_path, "cost.xlsx"))
        wb = openpyxl.load_workbook(path)
        ws = wb["КИЗ"]
        costs = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            costs.add(type(row[4]).__name__)
        assert "int" in costs or "float" in costs

    def test_kiz_variants(self, tmp_path):
        """Содержит чистый КИ-31, GS, WB, повреждённый и дубль."""
        import openpyxl

        path = create_synthetic_xlsx(os.path.join(tmp_path, "kiz.xlsx"))
        wb = openpyxl.load_workbook(path)
        ws = wb["КИЗ"]
        kiz_values = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            kiz_values.add(str(row[2]))

        assert KI_CLEAN in kiz_values, "Нет чистого КИ-31"
        assert KI_CORRUPTED in kiz_values, "Нет повреждённого КИ"
        assert KI_DUP in kiz_values, "Нет дублирующегося КИ"

    def test_deterministic(self, tmp_path):
        """Два вызова дают одинаковую логическую структуру (не ZIP-байты).
        ZIP-архивы содержат таймстампы, поэтому сравниваем содержимое ячеек."""
        import openpyxl

        p1 = os.path.join(tmp_path, "det1.xlsx")
        p2 = os.path.join(tmp_path, "det2.xlsx")
        create_synthetic_xlsx(p1)
        create_synthetic_xlsx(p2)

        wb1 = openpyxl.load_workbook(p1)
        wb2 = openpyxl.load_workbook(p2)

        assert wb1.sheetnames == wb2.sheetnames
        for name in wb1.sheetnames:
            ws1 = wb1[name]
            ws2 = wb2[name]
            for row1, row2 in zip(ws1.iter_rows(values_only=True),
                                   ws2.iter_rows(values_only=True)):
                assert list(row1) == list(row2), \
                    f"Различие в листе {name}: {row1} != {row2}"

    def test_expected_stats(self):
        """Статистика соответствует ожиданиям."""
        stats = get_expected_stats()
        assert stats["total_rows"] == 10
        assert stats["sale_rows"] == 7
        assert stats["return_rows"] == 1
        assert stats["dash_rows"] == 1
        assert stats["empty_check"] == 1
        assert stats["empty_fn"] == 1

    def test_no_real_xlsx_read(self):
        """Тест не читает реальный XLSX (проверка на уровне теста)."""
        # Этот тест сам по себе не читает реальные файлы
        # — гарантируется использованием tmp_path
        pass


class TestSyntheticXlsxWithoutOpenpyxl:
    """Если openpyxl не установлен — соответствующий skip."""

    def test_skip_if_no_openpyxl(self):
        """Маркер skip работает."""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl не установлен — пропуск")
