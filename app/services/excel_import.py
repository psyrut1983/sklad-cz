"""
excel_import — строгий разбор XLSX Wildberries (лист КИЗ).

Схема заголовков (10 колонок):
    № задания, Стикер, КИЗ, Номер чека, Стоимость, Валюта,
    Номер фискального накопителя, Дата, Тип операции, Признак продажи юрлицу

Поток:
  1. Открыть XLSX → проверить лист "КИЗ" и точный набор заголовков.
  2. Для каждой строки данных:
     a. Тип операции → только "Продажа" (trim); остальное → excluded.
     b. КИЗ → extract_ki (kiz_codec); невалидный → excluded.
     c. Номер чека, Номер фискального накопителя → непустые после trim.
     d. Стоимость → Decimal копейки; только RUB, >0.
     e. Дата → ISO "YYYY-MM-DD".
     f. Дубли KI-31 внутри файла → excluded.
  3. Итог: accepted + excluded + summary.

Полный КМ существует только в памяти вызова; наружу — только KI-31.
"""

from __future__ import annotations

import datetime
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any, BinaryIO, Sequence
import zipfile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from app.services.kiz_codec import extract_ki_or_none

# ═════════════════════════════════════════════════════════════════════════════
#  Константы
# ═════════════════════════════════════════════════════════════════════════════

HEADERS: list[str] = [
    "№ задания",
    "Стикер",
    "КИЗ",
    "Номер чека",
    "Стоимость",
    "Валюта",
    "Номер фискального накопителя",
    "Дата",
    "Тип операции",
    "Признак продажи юрлицу",
]

HEADER_KI_INDEX = 2       # КИЗ
HEADER_CHECK_INDEX = 3    # Номер чека
HEADER_COST_INDEX = 4     # Стоимость
HEADER_CURRENCY_INDEX = 5 # Валюта
HEADER_FN_INDEX = 6       # Номер фискального накопителя
HEADER_DATE_INDEX = 7     # Дата
HEADER_OP_INDEX = 8       # Тип операции

REQUIRED_CURRENCY = "RUB"

# ═════════════════════════════════════════════════════════════════════════════
#  DTO
# ═════════════════════════════════════════════════════════════════════════════


@dataclass(frozen=True)
class ExcludedRow:
    """Строка, исключённая из импорта."""
    row_index: int          # 1-based номер строки в листе (без заголовка)
    reason_code: str        # Стабильный машинный код причины
    message: str            # Человекочитаемое описание на русском (без КМ/КИЗ)


@dataclass(frozen=True)
class AcceptedRow:
    """Строка, прошедшая все проверки."""
    row_index: int
    ki: str                    # 31-символьный КИ-31 (без криптохвоста)
    check_number: str
    fn_number: str
    cost_kopecks: Decimal      # Стоимость в копейках
    date: str                  # ISO "YYYY-MM-DD"


@dataclass(frozen=True)
class ImportSummary:
    """Агрегированная статистика импорта."""
    total_rows: int = 0
    accepted: int = 0
    excluded: int = 0
    by_reason: dict[str, int] = field(default_factory=dict)
    accepted_submitted: int = 0
    accepted_failed: int = 0


@dataclass(frozen=True)
class ImportResult:
    """Результат разбора XLSX."""
    accepted: list[AcceptedRow] = field(default_factory=list)
    excluded: list[ExcludedRow] = field(default_factory=list)
    summary: ImportSummary = field(default_factory=ImportSummary)


# ═════════════════════════════════════════════════════════════════════════════
#  Ошибки уровня файла
# ═════════════════════════════════════════════════════════════════════════════


class FileImportError(ValueError):
    """Файл не удалось разобрать (невалидная структура)."""
    def __init__(self, message: str) -> None:
        super().__init__(message)


class MissingSheetError(FileImportError):
    """Нет листа КИЗ."""
    def __init__(self) -> None:
        super().__init__("Не найден лист «КИЗ»")


class InvalidHeadersError(FileImportError):
    """Заголовки не соответствуют ожидаемому набору."""
    def __init__(self, detail: str) -> None:
        super().__init__(f"Неверные заголовки: {detail}")


# ═════════════════════════════════════════════════════════════════════════════
#  Нормализация стоимости
# ═════════════════════════════════════════════════════════════════════════════


def _normalize_cost(value: Any) -> Decimal | None:
    """
    Нормализует стоимость в копейки Decimal.

    Принимает: int, Decimal, str (с точкой или запятой).
    Не принимает: float, bool, строки с >2 знаками после запятой.

    Возвращает Decimal(копейки) или None при ошибке.
    """
    # None и пустая строка
    if value is None:
        return None
    if isinstance(value, str) and value.strip() == "":
        return None

    # bool — reject (не является числом)
    if isinstance(value, bool):
        return None

    # float — принимаем только конечные, через Decimal(str()) для точности
    if isinstance(value, float):
        import math
        if math.isnan(value) or math.isinf(value):
            return None
        rubles = Decimal(str(value))
    elif isinstance(value, str):
        s = value.strip().replace(",", ".")
        if not s:
            return None
        try:
            rubles = Decimal(s)
        except InvalidOperation:
            return None
    elif isinstance(value, int):
        rubles = Decimal(str(value))
    elif isinstance(value, Decimal):
        rubles = value
    else:
        return None

    # <= 0
    if rubles <= 0:
        return None

    # >2 знака после запятой — отклоняем
    if rubles.as_tuple().exponent < -2:
        return None

    # Переводим рубли в копейки: умножаем на 100
    kopecks = (rubles * 100).to_integral_value()
    return kopecks


# ═════════════════════════════════════════════════════════════════════════════
#  Нормализация даты
# ═════════════════════════════════════════════════════════════════════════════

def _normalize_date(value: Any) -> str | None:
    """
    Нормализует дату в ISO "YYYY-MM-DD".

    Принимает: openpyxl date/datetime, строку "YYYY-MM-DD" или "DD.MM.YYYY",
    Excel serial number (через openpyxl.utils.datetime.from_excel).

    Возвращает строку "YYYY-MM-DD" или None.
    """
    if value is None:
        return None

    # datetime с таймзоной — не принимаем
    if isinstance(value, datetime.datetime):
        if value.tzinfo is not None:
            return None
        return value.strftime("%Y-%m-%d")

    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")

    # str
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        # ISO "YYYY-MM-DD"
        if len(s) == 10 and s[4] == "-" and s[7] == "-":
            try:
                datetime.date.fromisoformat(s)
                return s
            except (ValueError, TypeError):
                pass
        # Russian "DD.MM.YYYY"
        if len(s) == 10 and s[2] == "." and s[5] == ".":
            try:
                dt = datetime.datetime.strptime(s, "%d.%m.%Y")
                return dt.strftime("%Y-%m-%d")
            except (ValueError, TypeError):
                pass
        return None

    # int/float — Excel serial через openpyxl (без ручного int-truncate)
    if isinstance(value, (int, float)):
        if isinstance(value, bool):
            return None
        try:
            dt = openpyxl.utils.datetime.from_excel(value)
            return dt.strftime("%Y-%m-%d")
        except (ValueError, TypeError, OverflowError):
            return None

    return None


# ═════════════════════════════════════════════════════════════════════════════
#  Валидация заголовков
# ═════════════════════════════════════════════════════════════════════════════


def _validate_headers(actual: Sequence[str]) -> None:
    """
    Проверяет, что фактические заголовки совпадают с эталонными.

    Поднимает InvalidHeadersError при:
    - Неверном количестве колонок
    - Отсутствующих, лишних, дублированных заголовках
    - Несовпадении порядка
    """
    if len(actual) != len(HEADERS):
        raise InvalidHeadersError(
            f"ожидалось {len(HEADERS)} колонок, получено {len(actual)}"
        )

    for i, (expected, got) in enumerate(zip(HEADERS, actual)):
        if expected != got:
            raise InvalidHeadersError(
                f"колонка {i + 1}: ожидалось «{expected}», получено «{got}»"
            )


# ═════════════════════════════════════════════════════════════════════════════
#  Внутренние утилиты
# ═════════════════════════════════════════════════════════════════════════════


def _cell_str(cells: list[Any], idx: int) -> str:
    """Безопасное извлечение строки из ячейки с trim."""
    if idx >= len(cells):
        return ""
    val = cells[idx]
    if val is None:
        return ""
    if isinstance(val, str):
        return val.strip()
    if isinstance(val, bool):
        return ""
    return str(val).strip()


# ═════════════════════════════════════════════════════════════════════════════
#  Основной парсер
# ═════════════════════════════════════════════════════════════════════════════


def parse_xlsx(path: str | BinaryIO) -> ImportResult:
    """
    Разбирает XLSX Wildberries и возвращает структурированный результат.

    Параметры
    ---------
    path: str | BinaryIO
        Путь к файлу или открытый бинарный поток.

    Возвращает
    ----------
    ImportResult с accepted, excluded и summary.

    Исключения
    ----------
    FileImportError — при невалидной структуре файла.
    """
    wb = None
    try:
        try:
            if isinstance(path, str):
                wb = openpyxl.load_workbook(path, data_only=True)
            else:
                wb = openpyxl.load_workbook(path, data_only=True)
        except (zipfile.BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise FileImportError(
                "Не удалось открыть файл: повреждённый XLSX") from exc

        # ── Лист ───────────────────────────────────────────────────────────
        if "КИЗ" not in wb.sheetnames:
            raise MissingSheetError()

        ws = wb["КИЗ"]

        # ── Заголовки ──────────────────────────────────────────────────────
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        header_strs = [str(c) if c is not None else "" for c in header_row]
        _validate_headers(header_strs)

        # ── Данные ─────────────────────────────────────────────────────────
        accepted: list[AcceptedRow] = []
        excluded: list[ExcludedRow] = []
        seen_ki: set[str] = set()

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data_row_index = row_idx - 1  # 1-based, excluding header

            cells = list(row)

            # Полностью пустая строка (все ячейки None, пустая строка или bool)
            # — пропускаем, не считаем в summary
            if all(c is None or (isinstance(c, str) and c.strip() == "") or isinstance(c, bool) for c in cells):
                continue

            # ── Тип операции ───────────────────────────────────────────────
            raw_op = _cell_str(cells, HEADER_OP_INDEX)

            if raw_op != "Продажа":
                reason = _exclude_op_reason(raw_op)
                excluded.append(ExcludedRow(
                    row_index=data_row_index,
                    reason_code=reason,
                    message=_exclude_op_message(reason),
                ))
                continue

            # ── КИЗ ────────────────────────────────────────────────────────
            raw_kiz = _cell_str(cells, HEADER_KI_INDEX)
            if not raw_kiz:
                excluded.append(ExcludedRow(
                    row_index=data_row_index,
                    reason_code="empty_kiz",
                    message="КИЗ не заполнен",
                ))
                continue

            ki = extract_ki_or_none(raw_kiz)
            if ki is None:
                excluded.append(ExcludedRow(
                    row_index=data_row_index,
                    reason_code="invalid_kiz",
                    message="КИЗ не удалось преобразовать в КИ-31",
                ))
                continue

            # ── Номер чека ─────────────────────────────────────────────────
            check = _cell_str(cells, HEADER_CHECK_INDEX)
            if not check:
                excluded.append(ExcludedRow(
                    row_index=data_row_index,
                    reason_code="empty_check",
                    message="Номер чека не заполнен",
                ))
                continue

            # ── Номер фискального накопителя ───────────────────────────────
            fn = _cell_str(cells, HEADER_FN_INDEX)
            if not fn:
                excluded.append(ExcludedRow(
                    row_index=data_row_index,
                    reason_code="empty_fn",
                    message="Номер фискального накопителя не заполнен",
                ))
                continue

            # ── Стоимость ──────────────────────────────────────────────────
            raw_cost = cells[HEADER_COST_INDEX] if len(cells) > HEADER_COST_INDEX else None
            cost_kopecks = _normalize_cost(raw_cost)
            if cost_kopecks is None:
                excluded.append(ExcludedRow(
                    row_index=data_row_index,
                    reason_code="invalid_cost",
                    message="Стоимость не удалось нормализовать (только RUB, >0, не более 2 знаков после запятой)",
                ))
                continue

            # ── Валюта ─────────────────────────────────────────────────────
            raw_currency = _cell_str(cells, HEADER_CURRENCY_INDEX)
            if not raw_currency:
                excluded.append(ExcludedRow(
                    row_index=data_row_index,
                    reason_code="missing_currency",
                    message="Валюта не указана",
                ))
                continue

            if raw_currency.upper() != REQUIRED_CURRENCY:
                excluded.append(ExcludedRow(
                    row_index=data_row_index,
                    reason_code="unsupported_currency",
                    message="Поддерживается только RUB",
                ))
                continue

            # ── Дата ───────────────────────────────────────────────────────
            raw_date = cells[HEADER_DATE_INDEX] if len(cells) > HEADER_DATE_INDEX else None
            date_str = _normalize_date(raw_date)
            if date_str is None:
                excluded.append(ExcludedRow(
                    row_index=data_row_index,
                    reason_code="invalid_date",
                    message="Дату не удалось нормализовать",
                ))
                continue

            # ── Дубликат КИ ────────────────────────────────────────────────
            if ki in seen_ki:
                excluded.append(ExcludedRow(
                    row_index=data_row_index,
                    reason_code="duplicate_in_file",
                    message="Дубликат КИ-31 в файле (первое вхождение оставлено)",
                ))
                continue
            seen_ki.add(ki)

            # ── Принято ────────────────────────────────────────────────────
            accepted.append(AcceptedRow(
                row_index=data_row_index,
                ki=ki,
                check_number=check,
                fn_number=fn,
                cost_kopecks=cost_kopecks,
                date=date_str,
            ))

        # ── Сводка ─────────────────────────────────────────────────────────
        by_reason: dict[str, int] = {}
        for ex in excluded:
            by_reason[ex.reason_code] = by_reason.get(ex.reason_code, 0) + 1

        total = len(accepted) + len(excluded)
        summary = ImportSummary(
            total_rows=total,
            accepted=len(accepted),
            excluded=len(excluded),
            by_reason=by_reason,
        )

        return ImportResult(
            accepted=accepted,
            excluded=excluded,
            summary=summary,
        )

    finally:
        if wb is not None:
            wb.close()


# ═════════════════════════════════════════════════════════════════════════════
#  Причины исключения по типу операции
# ═════════════════════════════════════════════════════════════════════════════


def _exclude_op_reason(op: str) -> str:
    if op == "Возврат":
        return "return_operation"
    if op == "-":
        return "dash_operation"
    if op == "":
        return "empty_operation"
    return "unsupported_operation"


def _exclude_op_message(reason: str) -> str:
    messages = {
        "return_operation": "Операция «Возврат» не поддерживается",
        "dash_operation": "Строка без операции не поддерживается",
        "empty_operation": "Тип операции не указан",
        "unsupported_operation": "Тип операции не поддерживается",
    }
    return messages.get(reason, "Тип операции не поддерживается")
